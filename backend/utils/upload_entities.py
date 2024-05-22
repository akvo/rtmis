import os
import pandas as pd
from django.db.models import Q
from typing import List
from api.v1.v1_profile.models import (
    Administration,
    Levels,
    Entity,
    EntityData
)
import openpyxl
from utils.storage import upload


def generate_list_of_entities(
    file_path: str, entity_ids: List[int] = [], adm_id: int = None
):
    file_path = "./tmp/{0}".format(file_path.replace("/", "_"))
    if os.path.exists(file_path):
        os.remove(file_path)

    levels = Levels.objects.order_by("level").values("name")
    writer = pd.ExcelWriter(file_path, engine="xlsxwriter")
    entity_filter = Entity.objects.all()
    if entity_ids:
        entity_filter = Entity.objects.filter(id__in=entity_ids)
    for entity in entity_filter:
        entities = []
        filter_entity_data = EntityData.objects.filter(
            entity=entity,
        )
        if adm_id:
            administration = Administration.objects.get(id=adm_id)
            if administration.path:
                administration_path = (
                    administration.path + str(administration.id) + "."
                )
            else:
                administration_path = str(administration.id) + "."
            filter_entity_data = filter_entity_data.filter(
                administration__path__startswith=administration_path
            )
            # includes the administration itself
            filter_entity_data |= EntityData.objects.filter(
                administration=administration
            )
        for entity_data in filter_entity_data:
            administrations = entity_data.administration.full_path_name.split(
                "|"
            )
            entity_object = {
                "Name": entity_data.name,
                "Code": entity_data.code,
            }
            for i, level in enumerate(levels):
                if i < len(administrations):
                    entity_object[levels[i]["name"]] = administrations[i]
                else:
                    entity_object[levels[i]["name"]] = ""
            entities.append(entity_object)
        df = pd.DataFrame(entities)
        df.to_excel(writer, sheet_name=entity.name, index=False)
    writer.save()
    url = upload(file=file_path, folder="download_entities")
    return url


def validate_entity_data(filename: str):
    errors = []
    last_level = Levels.objects.all().order_by("level").last()
    xl = pd.ExcelFile(filename)
    wb = openpyxl.load_workbook(filename)
    for sheet in xl.sheet_names:
        check_sheet = wb[sheet]
        # skip empty sheets
        if all(cell.value is None for row
               in check_sheet.iter_rows() for cell in row):
            continue
        entity = Entity.objects.filter(name=sheet).first()
        if not entity:
            continue
        df = pd.read_excel(filename, sheet_name=entity.name)
        # remove rows with empty Name
        df = df.dropna(subset=["Name"])
        # remove exact duplicate rows
        df = df.drop_duplicates()
        if df.shape[0] == 0:
            errors.append({
                "sheet": entity.name,
                "message": "Empty data",
            })
        for index, row in df.iterrows():
            adm_names = []
            failed = False
            administration = None
            higher_level = None
            for level in Levels.objects.all().order_by("level"):
                if row[level.name] != row[level.name]:
                    previous_level = Levels.objects.filter(
                        level=level.level - 1
                    ).first()
                    if not higher_level:
                        higher_level = previous_level
                else:
                    row_value = row[level.name]
                    adm_names += [row_value]
                    administration = Administration.objects.filter(
                        parent=administration,
                        name=row_value,
                        level=level
                    ).first()
                    if not administration:
                        failed = True
                        continue
            if failed:
                adm_names = " - ".join(adm_names)
                errors.append({
                    "sheet": entity.name,
                    "row": index + 2,
                    "message": f"Invalid Administration for {adm_names}",
                })
            else:
                if level == last_level:
                    # skip if the entity data already exists
                    entity_name = row["Name"]
                    entity_data = EntityData.objects.filter(
                        Q(name__iexact=entity_name),
                        entity=entity,
                        administration=administration,
                    ).first()
                    if not entity_data:
                        code = row["Code"] if "Code" in df.columns else None
                        if code != code:
                            code = None
                        EntityData.objects.create(
                            name=entity_name,
                            code=code,
                            entity=entity,
                            administration=administration,
                        )
    return errors


def validate_entity_file(filename: str):
    xl = pd.ExcelFile(filename)
    wb = openpyxl.load_workbook(filename)
    sheet_names = xl.sheet_names
    errors = []
    # check if the sheet names are correct
    for sheet in sheet_names:
        check_sheet = wb[sheet]
        # skip empty sheets
        if all(cell.value is None for row
               in check_sheet.iter_rows() for cell in row):
            continue
        entity = Entity.objects.filter(name=sheet).first()
        if not entity:
            errors.append({
                "sheet": sheet,
                "row": 1,
                "message": f"Entity of {sheet} not found",
            })
        else:
            # check if the columns are correct
            df = pd.read_excel(filename, sheet_name=sheet)
            required_columns = Levels.objects.all().values_list(
                "name", flat=True
            )
            required_columns = list(required_columns) + ["Name", "Code"]
            for column in df.columns:
                if column not in required_columns:
                    errors.append({
                        "sheet": sheet,
                        "row": 1,
                        "message": f"Level {column} not found",
                    })
            if "Name" not in df.columns:
                errors.append({
                    "sheet": sheet,
                    "row": 1,
                    "message": "Name column not found",
                })
    return errors
